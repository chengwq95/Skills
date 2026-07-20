#!/usr/bin/env python3
"""Build a three-sheet e-commerce comparison workbook.

Usage: python build_price_excel.py data.json output.xlsx

Input example:
{
  "notes": ["Captured: 2026-07-20 10:00 MYT", "Delivery: Kuala Lumpur"],
  "rows": [{
    "product_id": "ssd-1tb", "name": "Example SSD", "platform": "Shopee",
    "shop": "Example Official Store", "spec": "1TB / blue", "currency": "MYR",
    "price": 448.04, "shipping": 0, "tax": 0, "discount": 0,
    "comparison_price": 448.04, "stock": "In stock", "price_type": "variant-selected",
    "note": "Public product-page price", "url": "https://example.invalid/item"
  }],
  "summary": [{"product_id": "ssd-1tb", "spec": "1TB / blue",
               "prices": {"Shopee": 448.04, "Lazada": 459.00},
               "currency": "MYR", "note": "In-stock offers only"}]
}

`summary` is optional. When omitted, it is calculated from available rows grouped by
product_id/name + spec + currency. The legacy fields name/platform/shop/spec/price/stock/url
remain sufficient for detail rows, but unknown shipping/tax/discount are not silently treated
as zero and therefore cannot produce an automatic winner.
"""
import argparse
import json
from collections import OrderedDict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PALETTE = ["FDECEA", "EAF1FB", "EAF7EA", "FFF6E5", "F3EAFB", "EAF7F7"]
REQUIRED = ("name", "platform", "shop", "spec", "stock", "url")


def number(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def available(stock):
    text = str(stock).lower()
    return not any(word in text for word in ("out of stock", "缺货", "unavailable", "sold out"))


def comparison_value(row):
    if number(row.get("comparison_price")):
        return row["comparison_price"]
    if number(row.get("price")) and all(number(row.get(key)) for key in ("shipping", "tax", "discount")):
        return row["price"] + row["shipping"] + row["tax"] - row["discount"]
    return None


def validate(data):
    if not isinstance(data, dict) or not isinstance(data.get("rows"), list):
        raise ValueError("Input must be an object with a rows array.")
    for index, row in enumerate(data["rows"], 1):
        if not isinstance(row, dict):
            raise ValueError(f"rows[{index}] must be an object.")
        missing = [field for field in REQUIRED if not row.get(field)]
        if missing:
            raise ValueError(f"rows[{index}] missing: {', '.join(missing)}")
        if row.get("price") is not None and not number(row["price"]):
            raise ValueError(f"rows[{index}].price must be numeric or null.")


def generated_summary(rows):
    groups = OrderedDict()
    for row in rows:
        key = (row.get("product_id") or row["name"], row["spec"], row.get("currency", "unknown"))
        groups.setdefault(key, []).append(row)
    result = []
    for (product_id, spec, currency), offers in groups.items():
        prices = OrderedDict()
        for offer in offers:
            value = comparison_value(offer)
            if value is None or not available(offer["stock"]):
                continue
            platform = offer["platform"]
            if platform not in prices or value < prices[platform]:
                prices[platform] = value
        unresolved = any(available(offer["stock"]) and offer.get("price") is not None and comparison_value(offer) is None
                         for offer in offers)
        note = "Automatically calculated from in-stock comparable offers"
        if unresolved:
            note += "; some item-price offers excluded because shipping/tax/discount is unknown"
        result.append({"product_id": product_id, "spec": spec, "currency": currency, "prices": prices, "note": note})
    return result


def main(data_path, out_path):
    with Path(data_path).open(encoding="utf-8") as source:
        data = json.load(source)
    validate(data)
    rows = data["rows"]
    summaries = data.get("summary") or generated_summary(rows)
    notes = data.get("notes", [])

    wb = openpyxl.Workbook()
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    body_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, count):
        for cell in ws[1][:count]:
            cell.font, cell.fill, cell.border = header_font, header_fill, border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(count)}1"

    detail = wb.active
    detail.title = "比价明细"
    headers = ["序号", "比较键", "商品名称", "平台", "店铺", "规格", "币种", "商品价", "运费", "税费", "折扣", "可比价格", "价格性质", "库存状态", "备注", "商品链接"]
    detail.append(headers)
    style_header(detail, len(headers))
    platforms = []
    for i, row in enumerate(rows, 1):
        platform = row["platform"]
        if platform not in platforms:
            platforms.append(platform)
        detail.append([i, row.get("product_id", row["name"]), row["name"], platform, row["shop"], row["spec"],
                       row.get("currency", "unknown"), row.get("price"), row.get("shipping"), row.get("tax"),
                       row.get("discount"), comparison_value(row), row.get("price_type", "unknown"), row["stock"],
                       row.get("note", ""), row["url"]])
        fill = PatternFill("solid", fgColor=PALETTE[platforms.index(platform) % len(PALETTE)])
        for cell in detail[detail.max_row]:
            cell.font, cell.border, cell.fill = body_font, border, fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for column in (8, 9, 10, 11, 12):
            cell = detail.cell(detail.max_row, column)
            if number(cell.value):
                cell.number_format = "#,##0.00"
            elif cell.value is None and column in (8, 12):
                cell.value = "—"
    widths = [6, 20, 36, 14, 28, 26, 10, 12, 12, 12, 12, 14, 18, 18, 30, 46]
    for column, width in enumerate(widths, 1):
        detail.column_dimensions[get_column_letter(column)].width = width

    summary = wb.create_sheet("比价汇总")
    platform_columns = []
    for item in summaries:
        for platform in item.get("prices", {}):
            if platform not in platform_columns:
                platform_columns.append(platform)
    summary.append(["比较键", "规格", "币种"] + [f"{p} 最低可比价" for p in platform_columns] + ["当前最优平台", "备注"])
    style_header(summary, 5 + len(platform_columns))
    for item in summaries:
        prices = item.get("prices", {})
        valid = {platform: value for platform, value in prices.items() if number(value)}
        winner = "需人工确认"
        if valid and item.get("currency", "unknown") != "unknown":
            lowest = min(valid.values())
            tied = [p for p, v in valid.items() if v == lowest]
            winner = "持平：" + "/".join(tied) if len(tied) > 1 else tied[0]
        summary.append([item.get("product_id", item.get("spec", "unknown")), item.get("spec", ""), item.get("currency", "unknown")] +
                       [prices.get(p) for p in platform_columns] + [winner, item.get("note", "")])
        for cell in summary[summary.max_row]:
            cell.font, cell.border = body_font, border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for column in range(4, 4 + len(platform_columns)):
            cell = summary.cell(summary.max_row, column)
            if number(cell.value):
                cell.number_format = "#,##0.00"
    for column, width in enumerate([20, 28, 10] + [18] * len(platform_columns) + [22, 48], 1):
        summary.column_dimensions[get_column_letter(column)].width = width

    notes_sheet = wb.create_sheet("说明")
    notes_sheet.append(["数据采集说明"])
    notes_sheet.cell(1, 1).font = Font(name="Arial", bold=True, size=12)
    default_notes = [
        "可比价格 = 商品价 + 已知运费 + 已知税费 − 保证适用的折扣。空白字段表示未知，不应视为零。",
        "仅可对相同比较键、相同币种的现货报价作出当前最优结论。",
    ]
    for note in [*default_notes, *notes]:
        notes_sheet.append([str(note)])
        notes_sheet.cell(notes_sheet.max_row, 1).font = body_font
        notes_sheet.cell(notes_sheet.max_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    notes_sheet.column_dimensions["A"].width = 120

    wb.save(out_path)
    print(f"saved: {out_path} rows={len(rows)} summary={len(summaries)}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build a three-sheet e-commerce comparison workbook.")
    parser.add_argument("data_json")
    parser.add_argument("output_xlsx")
    args = parser.parse_args()
    main(args.data_json, args.output_xlsx)
